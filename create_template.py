"""
Script to create the Excel template for margin calculation.
Run this once to generate the positions_template.xlsx file.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def create_excel_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Positions"

    # Header row styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Define headers (adjust these based on ICE upload format)
    headers = [
        "Account",
        "Symbol",
        "Quantity",
        "Price",
        "Side",  # BUY/SELL
        "Product Type",
        "Calculated Margin"  # This will be populated by the script
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Add sample data for reference
    sample_data = [
        ["ACC001", "ES", 10, 4500.00, "BUY", "Future", ""],
        ["ACC001", "NQ", 5, 15000.00, "SELL", "Future", ""],
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Adjust column widths
    column_widths = [12, 12, 12, 12, 10, 15, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width

    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        ["ICE Margin Calculator - Instructions"],
        [""],
        ["1. Fill in your position data in the 'Positions' sheet"],
        ["2. Save the Excel file"],
        ["3. Click 'Calculate Margin' button in the GUI application"],
        ["4. Wait for the calculation to complete"],
        ["5. The margin result will appear in the 'Calculated Margin' column"],
        [""],
        ["Column Descriptions:"],
        ["- Account: Your account identifier"],
        ["- Symbol: Trading symbol (e.g., ES, NQ, CL)"],
        ["- Quantity: Number of contracts"],
        ["- Price: Entry price"],
        ["- Side: BUY or SELL"],
        ["- Product Type: Future, Option, etc."],
        ["- Calculated Margin: Auto-filled by script (do not edit manually)"],
        [""],
        ["Note: Make sure to adjust the template columns to match"],
        ["the exact format required by ICE margin calculator upload."],
    ]

    for row_num, instruction in enumerate(instructions, 1):
        ws_instructions.cell(row=row_num, column=1, value=instruction[0])
        ws_instructions.cell(row=row_num, column=1).font = Font(bold=(row_num in [1, 9]))

    ws_instructions.column_dimensions['A'].width = 70

    # Save the template
    filename = "positions_template.xlsx"
    wb.save(filename)
    print(f"✅ Excel template created: {filename}")
    print("Edit this file with your positions, then use the GUI to calculate margin.")

if __name__ == "__main__":
    create_excel_template()
