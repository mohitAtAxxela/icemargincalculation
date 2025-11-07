"""
Main margin calculator module with Excel read/write functionality.
This replaces the old run_margin.py with single-file processing.
"""

import time
import re
from pathlib import Path
from queue import Queue
from threading import Event, Lock, Thread
from typing import Callable, Optional
from playwright.sync_api import sync_playwright
import pyperclip
from openpyxl import load_workbook

# ---------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------
SESSION_FILE = "ice_session.json"
APP_URL = "https://ica.ice.com/ICA/Main"
EXCEL_FILE = "positions_template.xlsx"  # Your working Excel file
RESULT_CELL_ID = "#cell-1468"  # The cell ID where margin result appears
# ---------------------------------------------------------------------


def read_excel_file(excel_path):
    """Read the Excel file and return data info."""
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    # Count data rows (excluding header)
    data_rows = sum(
        1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row)
        if any(cell.value is not None for cell in row)
    )

    print(f"📊 Loaded Excel: {Path(excel_path).name}")
    print(f"   Found {data_rows} position(s)")

    wb.close()
    return data_rows


def write_margin_to_excel(excel_path, margin_result):
    """Write the calculated margin back to Excel."""
    try:
        wb = load_workbook(excel_path)
        ws = wb.active

        # Find the "Calculated Margin" column
        margin_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(1, col).value and "Margin" in str(ws.cell(1, col).value):
                margin_col = col
                break

        if not margin_col:
            print("⚠️ 'Calculated Margin' column not found. Adding to column G.")
            margin_col = 7
            ws.cell(1, margin_col, "Calculated Margin")

        # Write margin result to row 2 (first data row)
        ws.cell(2, margin_col, margin_result)

        wb.save(excel_path)
        wb.close()

        print(f"✅ Margin result written to Excel: {margin_result}")
        return True

    except Exception as e:
        print(f"❌ Error writing to Excel: {e}")
        return False


class _Task:
    """Internal helper representing work for the Playwright worker thread."""

    def __init__(self, fn: Callable, args: tuple, kwargs: dict):
        self.fn = fn
        self.args = args
        self.kwargs = kwargs
        self._event = Event()
        self.result = None
        self.exception: Optional[Exception] = None

    def set_result(self, value):
        self.result = value
        self._event.set()

    def set_exception(self, error: Exception):
        self.exception = error
        self._event.set()

    def wait(self):
        self._event.wait()


class BrowserSession:
    """Manage a long-lived Playwright browser/page on a dedicated worker thread."""

    def __init__(self):
        self._task_queue: "Queue[object]" = Queue()
        self._thread: Optional[Thread] = None
        self._thread_lock = Lock()
        self._reload_event = Event()
        self._stop_sentinel = object()

    def _ensure_worker(self):
        with self._thread_lock:
            if self._thread is None or not self._thread.is_alive():
                self._thread = Thread(
                    target=self._worker_loop, name="BrowserSessionWorker", daemon=True
                )
                self._thread.start()

    def run(self, fn: Callable, *args, **kwargs):
        """Execute ``fn`` with a ready Playwright page on the worker thread."""

        self._ensure_worker()
        task = _Task(fn, args, kwargs)
        self._task_queue.put(task)
        task.wait()

        if task.exception is not None:
            raise task.exception

        return task.result

    def mark_needs_reload(self):
        """Indicate that the page should reload before the next task."""

        self._reload_event.set()

    def close(self):
        """Stop the worker thread and release Playwright resources."""

        with self._thread_lock:
            if self._thread and self._thread.is_alive():
                self._task_queue.put(self._stop_sentinel)
                self._thread.join()
            self._thread = None

    def _worker_loop(self):
        playwright = None
        browser = None
        context = None
        page = None
        initialized = False

        try:
            while True:
                task = self._task_queue.get()

                if task is self._stop_sentinel:
                    self._task_queue.task_done()
                    break

                try:
                    if self._reload_event.is_set():
                        initialized = False
                        self._reload_event.clear()

                    if playwright is None:
                        playwright = sync_playwright().start()

                    if browser is None or not browser.is_connected():
                        browser = playwright.chromium.launch(
                            headless=False, slow_mo=150
                        )
                        context = None
                        page = None
                        initialized = False

                    if context is None:
                        context = browser.new_context(storage_state=SESSION_FILE)
                        page = None
                        initialized = False

                    if page is None or page.is_closed():
                        page = context.new_page()
                        initialized = False

                    if not initialized:
                        print("\n🌐 Opening ICE ICA application...")
                        page.goto(APP_URL, timeout=60000)
                        page.wait_for_load_state("networkidle")
                        initialized = True

                    result = task.fn(page, *task.args, **task.kwargs)
                    task.set_result(result)

                except Exception as exc:  # noqa: BLE001 - propagate original error
                    initialized = False
                    self._reload_event.set()
                    task.set_exception(exc)

                finally:
                    self._task_queue.task_done()

        finally:
            try:
                if page and not page.is_closed():
                    page.close()
                if context:
                    context.close()
                if browser and browser.is_connected():
                    browser.close()
                if playwright:
                    playwright.stop()
            finally:
                self._reload_event.clear()


browser_session = BrowserSession()


def _perform_margin_calculation(page, excel_path: Path):
    """Core Playwright automation that must run on the worker thread."""

    # Check and clear existing portfolios
    checkbox_locator = (
        page.get_by_role(
            "gridcell",
            name="Press Space to toggle row selection (unchecked) All Portfolios (1)",
        )
        .get_by_label("Press Space to toggle row")
        .first
    )

    if checkbox_locator.count() > 0:
        print("🗑️  Clearing existing portfolios...")
        checkbox_locator.check()
        page.get_by_role("button", name="Actions").first.click()
        page.get_by_role("button", name="Delete").click()
        page.get_by_text("Delete", exact=True).click()
        page.get_by_role(
            "columnheader",
            name="Press Space to toggle all rows selection (unchecked) Calculation ID",
        ).get_by_label("Press Space to toggle all").first.check()
        page.get_by_role("button", name="Actions").nth(1).click()
        page.get_by_role("button", name="Delete").nth(1).click()
        page.get_by_role("button", name="OK").click()
        time.sleep(2)
    else:
        print("✓ No existing portfolios to clear")

    # Navigate to Tools → Upload Trades
    print("\n📤 Uploading positions file...")
    page.get_by_role("menuitem", name="Tools").click()
    page.get_by_role("menuitem", name="Upload Trades").click()

    # Upload the Excel file
    page.get_by_role("button", name=re.compile("Select file", re.I)).set_input_files(
        str(excel_path)
    )
    page.get_by_role("button", name="Upload").click()

    # Wait for upload confirmation
    page.wait_for_selector("button:has-text('OK')", timeout=60000)
    page.get_by_role("button", name="OK").click()
    time.sleep(3)
    okButtonLocator = page.get_by_role("button", name="OK")
    if okButtonLocator.is_visible():
        okButtonLocator.click()
    print("✅ Upload completed")

    # Select all accounts and run calculation
    print("\n🧮 Running margin calculation...")
    page.locator(
        "input[aria-label*='Press Space to toggle row selection']"
    ).first.check()
    page.get_by_role("button", name="Run Analytics").click()
    page.get_by_role("tabpanel").filter(has_text="Run").get_by_role("button").nth(
        1
    ).click()

    # Wait for calculation to complete (fixed time)
    print("⏳ Waiting for calculation to complete (5 seconds)...")
    time.sleep(5)  # Results appear within 5 seconds
    print("✅ Calculation completed")

    # Get the margin result directly from the cell
    print("\n📋 Extracting margin result...")

    # Method 1: Try to get text directly from the cell
    # try:
    #     result_cell = page.locator(RESULT_CELL_ID)
    #     result_cell.wait_for(timeout=30000, state="visible")  # 30 second timeout
    #     copied_text = result_cell.inner_text().strip()
    #     print(f"✅ Margin extracted: {copied_text}")
    # except Exception as e:
    #     print(f"⚠️ Direct extraction failed: {e}")
    #     # Method 2: Fallback to clipboard method
    #     print("Trying clipboard method...")
    #     page.locator(RESULT_CELL_ID).click(button="right", timeout=60000)
    #     time.sleep(0.5)
    #     page.get_by_text("Copy").first.click(timeout=5000)
    #     time.sleep(1)
    #     copied_text = pyperclip.paste().strip()
    #     print(f"✅ Margin copied via clipboard: {copied_text}")

    # # Result will be shown in the modal
    # print(f"\n{'='*60}")
    # print(f"✅ SUCCESS! Margin: {copied_text}")
    # print(f"{'='*60}\n")

    return


def run_margin_calc(excel_path, session: Optional[BrowserSession] = None):
    """
    Main function to run ICE margin calculator.
    1. Uploads the Excel file to ICE
    2. Runs the calculation
    3. Copies the result
    4. Writes result back to Excel
    """
    excel_path = Path(excel_path).resolve()

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    # Verify session file exists
    if not Path(SESSION_FILE).exists():
        raise FileNotFoundError(
            f"Session file '{SESSION_FILE}' not found. Please run 'login_once.py' first."
        )

    print(f"\n{'='*60}")
    print(f"Starting Margin Calculation")
    print(f"{'='*60}")

    # Read Excel to show info
    read_excel_file(excel_path)

    session = session or browser_session

    try:
        session.run(_perform_margin_calculation, excel_path)
        return
    except Exception as e:
        print(f"\n❌ Error during calculation: {e}")
        session.mark_needs_reload()
        raise


if __name__ == "__main__":
    # Test run
    try:
        run_margin_calc(EXCEL_FILE)
        # print(f"Final result: {result}")
    except Exception as e:
        print(f"Failed: {e}")
