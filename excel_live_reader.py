"""
Live Excel reader using COM API.
Reads data directly from open Excel instance without requiring save.
"""
import os
from pathlib import Path
import win32com.client
from openpyxl import load_workbook


def get_excel_instance():
    """Get the running Excel application instance."""
    try:
        excel = win32com.client.GetActiveObject("Excel.Application")
        return excel
    except:
        return None


def find_open_workbook(excel, file_path):
    """Find an open workbook by file path."""
    file_path = str(Path(file_path).resolve())

    for wb in excel.Workbooks:
        wb_path = str(Path(wb.FullName).resolve())
        if wb_path.lower() == file_path.lower():
            return wb
    return None


def read_excel_live(excel_path):
    """
    Read Excel data from either:
    1. Open Excel instance (if file is open) - gets unsaved changes
    2. Saved file (if file is closed) - uses openpyxl

    Returns: (data_rows, source_type)
    """
    excel_path = Path(excel_path).resolve()

    # Try to get open Excel instance
    excel = get_excel_instance()

    if excel:
        # Check if this specific file is open
        wb = find_open_workbook(excel, excel_path)

        if wb:
            print(f"📊 Reading from OPEN Excel (live data, unsaved changes included)")
            ws = wb.Worksheets(1)  # First sheet

            # Count data rows
            data_rows = 0
            row = 2  # Start from row 2 (skip header)

            while True:
                # Check if any cell in the row has data
                has_data = False
                for col in range(1, 8):  # Check first 7 columns
                    cell_value = ws.Cells(row, col).Value
                    if cell_value is not None and str(cell_value).strip() != "":
                        has_data = True
                        break

                if has_data:
                    data_rows += 1
                    row += 1
                else:
                    break

            print(f"   Found {data_rows} position(s) in open Excel")
            return data_rows, "live"

    # Fallback: Read from saved file
    print(f"📊 Reading from SAVED Excel file (file must be saved)")

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    # Count data rows (excluding header)
    data_rows = sum(1 for row in ws.iter_rows(min_row=2, max_row=ws.max_row)
                    if any(cell.value is not None for cell in row))

    print(f"   Found {data_rows} position(s) in saved file")
    wb.close()

    return data_rows, "saved"


def write_margin_to_excel_live(excel_path, margin_result):
    """
    Write margin result back to Excel.
    Works with both open and closed files.
    """
    excel_path = Path(excel_path).resolve()

    # Try to write to open Excel instance
    excel = get_excel_instance()

    if excel:
        wb = find_open_workbook(excel, excel_path)

        if wb:
            print(f"💾 Writing result to OPEN Excel (live update)")
            ws = wb.Worksheets(1)

            # Find the "Calculated Margin" column
            margin_col = None
            for col in range(1, 20):
                header = ws.Cells(1, col).Value
                if header and "Margin" in str(header):
                    margin_col = col
                    break

            if not margin_col:
                print("⚠️ Warning: 'Calculated Margin' column not found. Adding to column G.")
                margin_col = 7
                ws.Cells(1, margin_col).Value = "Calculated Margin"

            # Write margin result to row 2 (first data row)
            ws.Cells(2, margin_col).Value = margin_result

            # Add timestamp
            from datetime import datetime
            ws.Cells(2, margin_col + 1).Value = f"Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

            # Optional: Auto-save the workbook
            # wb.Save()  # Uncomment if you want auto-save

            print(f"✅ Margin result written to open Excel: {margin_result}")
            print(f"⚠️ Remember to save the Excel file manually!")
            return True

    # Fallback: Write to closed file using openpyxl
    print(f"💾 Writing result to SAVED Excel file")

    try:
        from openpyxl import load_workbook
        from datetime import datetime

        wb = load_workbook(excel_path)
        ws = wb.active

        # Find the "Calculated Margin" column
        margin_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(1, col).value and "Margin" in str(ws.cell(1, col).value):
                margin_col = col
                break

        if not margin_col:
            print("⚠️ Warning: 'Calculated Margin' column not found. Adding to column G.")
            margin_col = 7
            ws.cell(1, margin_col, "Calculated Margin")

        # Write margin result to row 2 (first data row)
        ws.cell(2, margin_col, margin_result)

        # Add timestamp
        ws.cell(2, margin_col + 1, f"Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        wb.save(excel_path)
        wb.close()

        print(f"✅ Margin result written to file: {margin_result}")
        return True

    except Exception as e:
        print(f"❌ Error writing to Excel: {e}")
        return False


if __name__ == "__main__":
    # Test the live reader
    test_file = "positions_template.xlsx"

    print("Testing live Excel reader...")
    print("=" * 60)

    try:
        rows, source = read_excel_live(test_file)
        print(f"\nResult: {rows} rows from {source} source")

        print("\nTesting write...")
        write_margin_to_excel_live(test_file, "123,456.78 USD")

    except Exception as e:
        print(f"Error: {e}")
